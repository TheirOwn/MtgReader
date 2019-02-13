# -*- coding: utf-8 -*-
"""
Created on Tue Feb  5 14:34:54 2019

@author: Nicholas
"""

from openpyxl import Workbook
from openpyxl.styles import colors, Alignment
from openpyxl.styles import Font

class Card:
    def __init__(self, name):
        #other tags: colors, colorIdentity
        self.name = name
        self.vipList = ["manaCost", "text", "subtypes", "types", "supertypes",
                        "power", "toughness", "legalities"]
        self.tags = dict()
        for key in self.vipList:
            self.tags[key] = ""
        self.ignorelist = ["layout", "tcgplayerProductId", "scryfallId",
                           "foreignData", "rulings","name", "tcgplayerPurchaseUrl", "type",
                           "uuid","convertedManaCost","starter", "faceConvertedManaCost",
                           "frameEffect", "isReserved","printings","colors","colorIdentity"]
    def addTag(self, tag, body):
        self.tags[tag] = body
    def search(self, key):
        if isinstance(key,str):
            key = key.lower()
        if key in self.name.lower():
            return True
        else:
            for keys, values in self.tags.items():
                if isinstance(values,str):
                    tvalue = values.lower()
                else:
                    tvalue=values
                if key in tvalue:
                    return True
        return False
    def getStats(self):
        if self.tags["power"] == "":
            self.tags["power"] = "*"
        if self.tags["toughness"] == "":
            self.tags["toughness"] = "*"
        return self.tags["power"] + "/" + self.tags["toughness"]
    def isCreature(self):
        return "Creature" in self.tags["types"] or "Summon" in self.tags["types"]
    def checkLegality(self,mode):
        """ legality text doesn't include names of illegal formats"""
        return self.tags["legalities"].find(mode.lower()) != -1
        
    def __repr__(self):
        return "Card " + self.name
    def __str__(self):
        ret = ""
        if self.isCreature():
            ret =  "Name: " + self.name + "\t\t" + self.tags["manaCost"]+"\n"+self.tags["supertypes"]\
                +" " + self.tags["types"] +"-"+self.tags["subtypes"]+"\n"+self.tags["text"]+\
                "\n"+self.getStats()
        else: 
            ret =  "Name: " + self.name + "\t\t" + self.tags["manaCost"]+"\n"+self.tags["supertypes"]\
                +" " + self.tags["types"] +"\n"+self.tags["text"]
        ret += "\n"
        for key,values in self.tags.items():
            if values == "" or key in self.ignorelist or key in self.vipList:
                continue
            ret += str(key) + ": " + str(values) + "\n"
        return ret

def readString(string):
    name = ""
    escape = False
    good = False
    counter = 0
    escchar = ""
    for c in string[1:]:
        counter += 1
        if escape:
            escape = False
        elif c == '"':
            good = True
            break
        if c == "\\":
            escape = True
            continue
        name += c
    if not good:
        print("BAD STRING")
        return "BAD STRING"
    #if name == "convertedManaCost" or name == "tcgplayerProductId" or name == "starter"\
    #   or name == "faceConvertedManaCost":
    #    return name,"",-1,string[counter+2:-1]
    if string.find("[", counter+1) ==-1 and string.find("{", counter+1) ==-1 \
        and string.find('"', counter+1) ==-1:
            return name,"",-1,string[counter+3:-1]
    
    
    secondstring = string.find('"', counter+1)
    bracket = string.find("]", counter+1) + string.find("}", counter+1)
    text = ""
    level = 0
    if secondstring > -1:
        for c in string[secondstring+1:]:
            if escape:
                escape = False
                if c == "n":
                    text += "\\"
            elif c == '"':
                good = True
                break
            if c == "\\":
                escape = True
                continue
            text += c
        level = -1
    elif bracket > -2:
        level = -1
    if string[-1] == "[":
        escchar = "],"
    elif string[-1] == "{":
        escchar = "},"
    #need to handle [] or {} in one line to remove level
    if not good:
        print("BAD STRING")
        return "BAD STRING"
    return name,escchar,level,text,
    
def printExcel2(cards,traits, wbname, searchtags, mode="commander"):
    #mode is only for legality
    wb = Workbook()
    ws = wb.active
    
    ws["A1"].value = "Nick's Card Searcher V1.0"
    ws["A1"].font = Font(bold=True,color=colors.BLUE,size=15)
    ws["A2"].value = "Keyword = "+ '"' + (' "').join(searchtags) + '"'
    ws.column_dimensions["A"].width = 30
    
    cell = ws["A4"]
    for trait in traits:
        cell.value = trait.title()
        if trait == "alltypes":
            cell.value = "Type"
        cell.font = Font(bold=True)
        cell = cell.offset(0,1)
    cell = ws["A5"]
    for spell in cards:
        for trait in traits:
            if trait == "alltypes":
                cell.value = ""
                if spell.tags["supertypes"] != "":
                    cell.value += spell.tags["supertypes"] + " "
                cell.value += spell.tags["types"]
                if spell.tags["subtypes"] != "":
                    cell.value +="- " + spell.tags["subtypes"]
            elif trait in spell.tags.keys():
                if trait == "power" or trait == "toughness":
                    if not spell.isCreature():
                        cell = cell.offset(0,1)
                        continue
                if trait == "tcgplayerPurchaseUrl" or (trait == "name" and "tcgplayerPurchaseUrl" in spell.tags.keys()):
                    ttrait = '=Hyperlink("'+spell.tags["tcgplayerPurchaseUrl"]+'","'+spell.tags["name"]+'")'
                    cell.font = Font(u='single', color=colors.BLUE)
                elif trait == "legalities":
                    cell.value = spell.checkLegality(mode)
                else:
                    try:
                        ttrait = int(float(spell.tags[trait]))
                    except:
                        ttrait = spell.tags[trait]
                        if r"\n" in ttrait:
                            cell.alignment = Alignment(wrapText=True,vertical="top")
                            ttrait = ttrait.replace("\\n","\n")
                            tcell = cell
                            while True:
                                try:
                                    tcell = tcell.offset(0,-1)
                                    tcell.alignment = Alignment(wrapText=True,vertical="center")
                                except:
                                    break
                cell.value = ttrait
            cell = cell.offset(0,1)
        cell = ws.cell(row = cell.row + 1, column = 1)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.1,135)
        if adjusted_width == 135:
            for cell in col:
                cell.alignment = Alignment(wrapText=True,vertical="top")
        if column == "A":
            adjusted_width = 30
        ws.column_dimensions[column].width = adjusted_width
    
    #wb.create_sheet("Card Format")
    #ws=wb["Card Format"]
    #cell = ws["A1"]
    
    #for spell in cards:
        
    
    
    wb.save(wbname)
    wb.close()
    
def main():
    user = ""
    searchtags = []
    print("Input a word or exact phrase to search for, or enter " +\
                     "\"start search\" to begin search. If you put a '~' before the word,"+\
                     "then that criteria will be optional, but will be marked as a high match " +\
                     "if it is present:\n")
    while user.lower() != "start search":
        if user != "":
            searchtags.append(user.lower())
        user = input("")
    file = r"Cards\AllCards.json"
    level = 0
    tag = ""
    body = ""
    cards = []
    highmatch = []
    coins = 100000
    escchar = ['],','']
    print("Loading cards", end="",flush=True)
    with open(file, "r", encoding="utf8") as f:
        f.readline(); #toss opening {
        for line in f:
            if coins % 50000 == 0:
                print(".",end="",flush=True)
            line = line.lstrip()[:-1]
            if len(line) == 0:
                print("DONE")
                break
            else:
                #level 0: entire card
                #level 1: A tag
                #level 2: Inside a multi tag
                if level == 0 and line[0] == '"': #name of the card
                    """beginning of a card"""
                    name,escchar[level],tlevel,body = readString(line)
                    cards.append(Card(name))
                    level += 1
                elif level == 2 and line != escchar[1] and line[0] == '"':
                    if body != "":
                        body += ", "
                    bodyt,tlevel,escchart,bodyt2 = readString(line) #body is returned in 0 if its only string
                    body += bodyt + bodyt2
                    #if escchart != "":
                     #   escchar[level] = escchart
                elif level == 2 and line == escchar[1]:#end of multiline tag
                    cards[-1].addTag(tag,body)
                    level -=1
                elif level == 1 and line != escchar[0] and line[0] == '"':#begintag
                    tag,escchart,tlevel,body = readString(line)
                    if escchart != "":
                        escchar[level] = escchart
                    if tlevel == -1:
                        cards[-1].addTag(tag,body)
                    else:
                        level +=1
                elif level == 1 and line == escchar[0]:
                    """end of the card"""
                    #print(cards[-1])
                    if coins == 0:
                        break
                    coins -=1
                    level -= 1
                    for keys in searchtags:
                        if not cards[-1].search(keys):
                            if keys[0] != "~" and keys[0] != "-":
                                del cards[-1]
                                break
                        elif keys[0] == "-":
                            del cards[-1]
                            break
                        elif keys[0] == "~":
                            highmatch.append(cards[-1])
        if not cards[-1].search(keys):
            if keys[0] != "~" and keys[0] != "-":
                del cards[-1]
        elif keys[0] == "-":
            del cards[-1]
        elif keys[0] == "~":
            if cards[-1] not in highmatch:
                highmatch.append(cards[-1])
                        
    f.close()
    print("Finished loading. Now printing")
    
    
    #save = []
    #for x in cards:
     #   if x.search("Merfolk"):
      #      save.append(x)
    mode = "commander"
    printExcel2(highmatch + cards,["name","manaCost", "alltypes","text","power",\
                           "toughness","colors","legalities"],"MTGResults.xlsx",\
                searchtags,mode)

if __name__  == "__main__":
    main()